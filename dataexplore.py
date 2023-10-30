import plotly.express as px
import pandas as pd
import openpyxl
from dash import Dash, dcc, html, Input, Output

file_path = 'dadosEletricidade.xlsx'

def extrair_dados(start_cell, end_cell, sheet):
    cell_range = []

    for row in sheet.iter_rows(min_row=int(start_cell[1:]), min_col=ord(start_cell[0]) - 64,
                            max_row=int(end_cell[1:]), max_col=ord(end_cell[0]) - 64,
                            values_only=True):
        for cell_value in row:
            cell_range.append(cell_value)

    return pd.Series(cell_range)


df = pd.DataFrame()

workbook = openpyxl.load_workbook(file_path)
sheets = workbook.sheetnames

df['Cidade'] = extrair_dados('B4', 'B648', workbook['2008'])

for i, sheet in enumerate(sheets):
    start_cell = 'Q4'
    end_cell = 'Q648'
    curr_sheet = workbook[sheet]
    df[sheet] = extrair_dados(start_cell, end_cell, curr_sheet)

df_melted = df.melt(id_vars='Cidade', var_name='Ano', value_name='MWh')
df_melted['Ano'] = df_melted['Ano'].astype(int)

app = Dash(__name__)
app.layout = html.Div(
    [html.Header("Dados Energéticos do Estado de São Paulo"),
     dcc.Dropdown(id = 'dpd', 
                  options=df['Cidade'], 
                  value='Campinas'),
    dcc.Graph(id = 'scattergraph')]
)

@app.callback(Output('scattergraph', 'figure'),
              Input('dpd', 'value'))

def sync_input(city_input):
    fig = px.line(df_melted[df_melted['Cidade'] == city_input],
                  x='Ano',
                  y='MWh',
                  color='Cidade',
                  labels={'MWh': 'Energia MWh'})

    fig.update_layout(
        title='Energia em MWh por ano (2008-2022)',
        xaxis_title='Ano',
        yaxis_title='Energia (MWh)',
    )

    return fig

if __name__ == '__main__':
    app.run_server(debug = False)